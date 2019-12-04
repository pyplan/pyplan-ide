#functions for use as:  import cubepy.factory as cp

import numpy as np
import cubepy
import numbers
from sys import platform
from cubepy.cube import kindToString, apply_op
import os
import csv

random = -123798
byVal = 1
byPos = 2
exact = 1
start_with = 2
end_with = 3
contain = 4




def cube(axes, values=None, broadcast=True,dtype=None):
    """Create a cube  object. 
    axes: list of axis of the cube
    values: optional, list of values of the cube. Can be other cubes for build a report.
    Ex.
        cp.cube([time])
        cp.cube([time,product])
        cp.cube([time,product],[10,20,30,40,50,60,70,80])
        cp.cube([time,product],cp.random)
        
        cp.cube([index_reports],[report_1,report_2])
    """
    if values is None:
        if not dtype is None:
            if dtype is str:
                return cubepy.Cube.full(axes,'', dtype='U25')
            elif kindToString(np.dtype(dtype).kind)=="string":
                return cubepy.Cube.full(axes,'', dtype=dtype)
        return cubepy.Cube.zeros(axes)
    else:
        if isinstance(values,list) or isinstance(values,np.ndarray) :
            if len(values)>0:
                if isinstance(values[0],cubepy.Cube):
                    #use stack
                    if isinstance(axes,list):
                        axes = axes[0]
                    return cubepy.stack(values,axes,broadcast)
            return cubepy.Cube(axes,values,fillValues=True, dtype=dtype)
        elif isinstance(values,numbers.Number) and values==random:
            theSize = [len(x) for x in axes]
            return cube(axes, np.random.randint(100, size=theSize), dtype=dtype)
        else:
            return cubepy.Cube.full(axes,values, dtype=dtype)


def index(name,values):
    """Create a index object.
    name: name for the index
    values: list of values of the index.
    Ex.
        cp.index("items",["Item 1","Item 2","Item 3"])
        cp.index("years",[2016,2017,2018])
    """
    if values is None:
        values = ["Item 1","Item 2","Item 3"]
    return cubepy.Index(name ,values)            


def find(param1, param2, compareType=1, caseSensitive = True):
    """
    param1: value or indexarray for compare
    param2: index compare to
    compareType: cp.exact=1, cp.start_with=2, cp.end_with=3, cp.contain=4  
    caseSensitive: able to differentiate between uppercase and lowercase (by default True)

    If param1 is a scalar (numeric or str) and param2 is an index:  return cube indexed by param2 with True on ocurrences of param2
        Ex. result = cp.apply("te", region, cp.end_with)
    If param1 is an index and param2 is an index too:  return cube indexed by param1 and param2 with True on ocurrences of param1 on param2
        Ex. result = cp.apply(subregion, region, cp.contain)

    """
    def _fn(item,value):
        if isinstance(item,str) == False:
            item = str(item)
        if isinstance(value,str) == False:
            value = str(value)
        if compareType==1:
            if caseSensitive:
                return item == value
            else:
                return item.lower() == value.lower()
        elif compareType==2:
            if caseSensitive:
                return item[:len(value)] == value
            else:
                return item[:len(value)].lower() == value.lower()                                    
        elif compareType==3:
            if caseSensitive:
                return item[-len(value):] == value
            else:
                return item[-len(value):].lower() == value.lower()                    
        elif compareType==4:
            if caseSensitive:
                return value in item
            else:
                return value.lower() in item.lower()

    if (isinstance(param1,str) or str(param1).isnumeric()) and isinstance(param2,cubepy.Index):
        vfn = np.vectorize(_fn)
        return cubepy.Cube([param2],vfn(param2.values,param1))

    if isinstance(param1,cubepy.Index) and isinstance(param2,cubepy.Index):
        _res = cubepy.Cube.full([param1,param2],False)
        rr=0
        for row in param1.values:
            cc=0
            for col in param2.values:
                _res.values[rr,cc] = _fn(col,row)
                cc+=1
            rr+=1
        return _res

def selectText(data, first = None, last = None ):
    """Returns a new cube with the text contained between the 'first' and 'last' characters of cube / index 'data'. Starts counting from 0.
            If 'first' character is ommited, it returns every character from the first character of 'data' to the 'last' character, inclusive.
            If 'last' character is ommited, it returns every character from "first" character of 'data', to the last character available for each cell.
    """
    if first is None:
        if last is None:
            sliced_data = apply( lambda x: x[:], data )
        else:
            sliced_data = apply( lambda x: x[:last], data )
    else:
        if last is None:
            sliced_data = apply( lambda x: x[first:], data )
        else:
            sliced_data = apply( lambda x: x[first:last], data )
    
    return sliced_data        
        

def apply(fn, param1, param2=None, start=None):
    """
    Apply functions to index and cubes. Multiple results can be obtained
    fn: function to apply
    param1: index or cube
    param2: index or cube
    start: scalar or cube
        Ex. 
            cp.apply(lambda x: x[:2] ,indexRegion): return new cube indexed by "indexRegion" and apply fn on each item
            cp.apply(lambda x: x*5 ,cubeSales): return new cube result of apply fn on all values of the cubeSales
            cp.apply( cp.addPeriods, start_proj , end_proj): Return new cube result of apply fn on "start_proj" with "end_proj"
            cp.apply(lambda x: x+1, indexYear, start=10) : Return new cube indexed by "indexYear", result of apply fn starting with scalar value "10"
            cp.apply(lambda x: x+1, indexYear, start=prices) : Return new cube indexed by "indexYear", result of apply fn starting with cube "prices"
    """
    if callable(fn):
        vfn = np.vectorize(fn)
        if param2 is None:
            if isinstance(param1,cubepy.Index):
                if start is None:
                    return cubepy.Cube([param1],vfn(param1.values)) 
                elif isinstance(start,cubepy.Cube):
                    values=[start.values]
                    numEls =len(param1)
                    for nn in range(numEls-1):
                        values.append(fn(values[nn]))
                    new_axes = start._axes.insert(param1, 0)
                    return cubepy.Cube(new_axes,values)
                else:
                    values=[start]
                    numEls =len(param1)
                    for nn in range(numEls-1):
                        values.append(fn(values[nn]))
                    return cubepy.Cube(param1,values)
            if isinstance(param1,cubepy.Cube):
                return param1.apply(fn)
        elif isinstance(param1,cubepy.Cube) and isinstance(param2,cubepy.Cube):
                return apply_op( param1 , param2 , vfn)
    return None


def max(cube1,cube2):
    """Return max value between two cubes
    """
    return (cube1 > cube2)*cube1 + (cube1 <= cube2) * cube2

def min(cube1,cube2):
    """Return min value between two cubes
    """
    return (cube1 > cube2)*cube2 + (cube1 <= cube2) * cube1

def sum(cube, axis=None, keep=None, group=None, sort_grp=True):
    """Sum of array elements over a given axis.
    :param axis: Axis or axes along which a sum is performed. The default (axis = None) is perform a sum
    over all the dimensions of the input array. axis may be negative, in which case it counts from the last
    to the first axis. If this is a tuple of ints, a sum is performed on multiple axes, instead of a single
    axis or all the axes as before.
    :return: new Cube instance or a scalar value
    """
    return cube.reduce(np.sum, axis, keep, group, sort_grp)    


def subscript(cube, index, value):
    """Filter cube1 using the index and the value. Return a new cube without the index1 dimension
        Ex.
            cp.subscript(nodo_ejemplo,index_para_ejemplo,"Item 1")
    """
    return cube[index==value]


def slice(cube, index, value):
    """Filter cube using the index and the value. Return a new cube without the index dimension
        Ex.
            cp.slice(nodo_ejemplo,index_para_ejemplo,2)
    """
    if isinstance(index,cubepy.Index):
        index = index.name
    if isinstance(value,str) or str(value).isnumeric():
        value =[value]
    return cube.take(index,value).squeeze()


def shift(cube, axis, n=1, cval=0):
    """Returns a cube with the axis shifted.
        Ex.
            cp.shift(nodo_ejemplo,index_para_ejemplo,1)
    """
    return cube.shift(axis,n,cval)


def subset(cube, indexName="new index"):
    """Returns a list of all the elements of the index for which cube is true. The function is used to create a new index that is a subset of an existing index.
        Ex. cp.subset(cantidades>0)
    """
    cond = cube>0
    values = cond.axes[0].values[cond.values]
    return index(indexName,values)


def aggregate(cube,mapCube,indexToRemove,targetIndex):
    """ Aggregates the values in Cube to generate the result indexed by  targetIndex.
        Map gives the value of targetIndex for each element of indexToRemove

        Example for aggregating time information into annual index the syntax is:
            cp.aggregate(cube, map, time, years )
    """
    grouping_index_mat = cubepy.Cube([targetIndex],targetIndex.values)
    mat_allocation = mapCube == grouping_index_mat
    return (cube * mat_allocation).sum(indexToRemove)


def cumulate(cube, index):
    """ TODO coment
    """
    pos=0
    tmpMat=cubepy.Cube.zeros(cube.axes)
    tmpInd=cubepy.Cube.zeros([index])
    for j in index.values:
        tmpInd.values[pos:pos+1]=1
        tmpMat = tmpMat + ((tmpInd * cube).sum(index))*(j==index)
        pos=pos+1
    return tmpMat


def cumProd(cube, index):
    """Return the cumulative product of elements along a given axis
        param cube: cube 
        param axis: axis name (str), index (int) or instance
        Ex:
            cp.cumProd(nodo,indice)
    """
    return cube.cumProd(index)


def irr(flow, time_index ):
    """Returns the Internal Rate of Return (IRR) of a series of periodic payments (negative values) and inflows (positive values). The IRR is the discount rate at which the Net Present Value (NPV) of the flows equals zero. 
        The variable flow must be indexed by time_index.

    If the cash flow never changes sign, cp.irr() has no solution and returns NAN (Not A Number).
    """
    import pandas as pd

    _cube_dimensions = index("flowdims",flow.dims ) 
    _rest_of_indexes_labels = subset( _cube_dimensions != time_index.name ).values 
    _rest_of_indexes = [flow.axis(xx) for xx in _rest_of_indexes_labels]

    _cube = None
    if len( _rest_of_indexes ) == 0: 
        _cube = np.irr( flow.values ) 
    else: 
        _cube = cube( _rest_of_indexes ) 
        _multivalues = [idx.values for idx in _rest_of_indexes] 

        _values = pd.MultiIndex.from_product( _multivalues ).values 

        for _item in _values: 
            _filter = []

            for _nn in range(len(_item)): 
                _filter.append( _rest_of_indexes[_nn].filter( [_item[_nn]] ) ) 

            _irr = np.irr( flow.filter( _filter ).squeeze().values ) 
            _cube.set_data( _filter, _irr )
    return _cube


def npv(rate, flow, time_index, offset = 1): 
    """"Returns the Net Present Value (NPV) of a cash flow with equally spaced periods. The flow parameter must contain a series of periodic payments (negative values) and inflows (positive values), indexed by time_index.
        The optional offset parameter especifies the offset of the first value relative to the current time period. By default, offset is set to 1, indicating that the first value is discounted as if it is one step in the future
    """
    _number_of_periods = time_index.pos + offset 
    _present_values = flow / ( 1 + rate ) ** _number_of_periods 
    _npv = _present_values.sum( axis = time_index ) 
    return _npv


def splitText(text,separator=",",resultIndex=None, indexName="new index"):
    """Split text into a list of substrings by each occurrence of separator.
        text:  string or cp.index used for split
        separator: string for use as separator
        resultIndex: optional, for return cp.cube indexes by this index
        indexName: optional, string name for new index
        Ex:
            cp.splitText("Uno Dos Tres Cuatro Cinco Seis", separator=" ") 
            cp.splitText(indice_para_split, separator="-") 
            cp.splitText("Uno Dos Tres Cuatro Cinco Seis"," ",result_index)  
            cp.splitText(indice_para_split, "-", result_index)  

    """
    if isinstance(text,str):
        _arr = text.split(separator)
        if resultIndex is None:
            return self.index(indexName,_arr)
        else:
            return self.cube(resultIndex,_arr)

    elif isinstance(text,cubepy.Index):
        _arr=[]
        for item in text.values:
            _tmpArr = str(item).split(separator)
            if not resultIndex is None:
                _tmpNpArray = np.array(_tmpArr)
                _tmpArr=[]
                for _nn in range(len(resultIndex)):
                    if (_nn<len(_tmpNpArray)):
                        _tmpArr.append(_tmpNpArray[_nn])
                    else:
                        _tmpArr.append("")

            _arr = _arr + _tmpArr
        _arrnp = np.array(_arr).flatten()
        
        if resultIndex is None:
            return index(indexName,_arrnp)
        else:
            return cube([text,resultIndex],_arrnp)


def changeIndex(cube, oldIndex, newIndex, compareMode=1,default=None):
    """Returns a new Cube instance with the axes changed,looking newIndex on oldIndex by value o by position.
        param Comparemode: cp.byVal or cp.byPos
        Ex:
            cp.changeIndex(nodo_original,indice_original,indice_cambiado,cp.byVal)
            cp.changeIndex(nodo_original,indice_original,indice_cambiado,cp.byPos)
    """
    return cube.change_axis(oldIndex,newIndex,compareMode,default=default)            


def copyIndex(cube, indexName="new index"):
    """Generate a cp.index with current unique values of the cube. 
    The cube must have only one dimension
    """
    import pandas as pd
    if cube.ndim>1:
        raise ValueError("The cube must have only one dimension")

    np_values = cube.values.flatten()
    seripandas = pd.Series(np_values)
    return index(indexName,seripandas.unique())


def cascadeVolume(demand,ranges,consumption_range_index=None):
    """TODO Comment
    Ex:
            cp.cascadeVolume(cantidades,limites_rango_consum)
            cp.cascadeVolume(cantidades,limites_rango_consum,"rangos_consumo")
            cp.cascadeVolume(cantidades,limites_rango_consum,rangos_consumo)
    """
    if consumption_range_index is None:
        if ranges.ndim==1:
            consumption_range_index = ranges.dims[0]
        else:
            raise ValueError("You must specify the consumption_range_index")
    elif isinstance(consumption_range_index,cubepy.Index):
        consumption_range_index = consumption_range_index.name


    vol_entre_lim = ranges - ranges.shift(consumption_range_index,-1,0)
    ConsumoBajoLimite = (ranges<demand)*vol_entre_lim
    # Forma de determinar el ultimo umbral de consumo
    ConsEnRangos = (demand>ranges)*1
    ultrango = (ConsEnRangos - ConsEnRangos.shift(consumption_range_index,-1,1))==-1
    ConsumoLimite = (demand - ConsumoBajoLimite.sum(consumption_range_index)) * ultrango
    return ConsumoBajoLimite + ConsumoLimite



def bandAllocation(demand,ranges,consumption_range_index=None):
    """TODO Comment
        Ex:
            cp.bandAllocation(cantidades,limites_rango_consum)
            cp.bandAllocation(cantidades,limites_rango_consum,"rangos_consumo")
            cp.bandAllocation(cantidades,limites_rango_consum,rangos_consumo)
    """

    if consumption_range_index is None:
        if ranges.ndim==1:
            consumption_range_index = ranges.dims[0]
        else:
            raise ValueError("You must specify the consumption_range_index")
    elif isinstance(consumption_range_index,cubepy.Index):
        consumption_range_index = consumption_range_index.name

    ConsEnRangos = (demand>ranges)*1
    ultrango = (ConsEnRangos - ConsEnRangos.shift(consumption_range_index,-1,1))==-1
    ConsumoLimite = demand * ultrango
    return ConsumoLimite


def dispatch(contract_vol, contract_price, contract_index, demand):
    """TODO comment
        Ex:
            cp.dispatch(contract_volumes,prices,contracts,demand)
    """
    order_index = index("order",range(len(contract_index)))
    contract_order = argsort(contract_price,contract_index)
    vol_by_dispatch_order = (contract_order==order_index)*contract_vol
    cum_vol = cumulate(vol_by_dispatch_order,order_index).sum(contract_index)
    return (cascadeVolume(demand,cum_vol,order_index)*(contract_order==order_index)).sum(order_index)


def argsort(cube,axis):
    """Return new cube with the position of values sorted by axis
        ex.
            cp.argsort(nodo_1,index_a)
    """
    return cube.argsort(axis)   


def iif(condition,truePart, falsePart=None):
    """Inline if. Evaluate condition and return truePart or fasePart
        ex. 
            cp.iif( producto=="Producto A", cantidades*1000 , cantidades)
            cp.iif( cantidades > 10, cantidades , 0)
    """
    if isinstance(condition,cubepy.Cube):
        _true  = condition * truePart
        if falsePart is None:
            return _true
        else:
            _false = (~condition) * falsePart
            return _true + _false
    elif isinstance(condition,bool):
        if condition:
            return truePart
        else:
            return falsePart
    return None    




def cubeFromPandas(dataframe, cubeIndexes, valueColumns, indexColumnHeaders=None,replaceByIndex=None):
    """Create new cp.cube, converting pandas to multidimensional data, according to the parameters
        dataframe: pandas dataframe
        cubeIndexes: objects cp.index for perform change index
        valueColumns: string with column name of the dataframe where contain the values
                        cp.index with columns names for convert colums to index
        indexColumnHeaders: (optional) column names in pandas to parse with cubeIndexes. Used if header on dataframe is not equal to index identifiers.
        replaceByIndex: (optional) replace index used in valueColumns for this index. (using changeindex)
    """
    import pandas as pd

    valueIndex=None
    if isinstance(valueColumns,cubepy.Index):
        valueIndex = valueColumns
        valueColumns = valueIndex.values
    elif isinstance(valueColumns,str):
        valueColumns = np.array([valueColumns])

    if indexColumnHeaders is None:
        indexColumnHeaders = [index.name for index in cubeIndexes]
        
    #create total index and index names

    _allindexes = cubeIndexes
    _allIndexNames = indexColumnHeaders[:]
    if not valueIndex is None:
        _allindexes.append(valueIndex)
        _allIndexNames.append("data_index")    

    #fill other columns for prevent melt error
    if isinstance(dataframe, pd.DataFrame):
        cols_not_in_df = [col for col in valueColumns if col not in dataframe.columns]
        for col in cols_not_in_df:
            dataframe[col] = np.nan


    _full = dataframe.reset_index().melt(id_vars=indexColumnHeaders,value_vars=valueColumns, var_name="data_index", value_name="data_value")

    if _full.size==0:
        _cube = cube(_allindexes, np.array([],dtype="O"))
    else:

        #sum for acum over duplicate records
        _full = _full.groupby(_allIndexNames, as_index=False).sum()
        _dtype = _full["data_value"].dtype
        
        _dataType = kindToString(_dtype.kind )
        if _dataType=="string":
            _full = _full[ (_full["data_value"]!="") & (_full['data_value'].notna()) ]
        else:
            _full = _full[ (_full["data_value"]!=0) & (_full['data_value'].notna()) ]


        _size = [len(x) for x in _allindexes]
        _emptyData = np.zeros(_size, dtype=_dtype )
        _cube = cube(_allindexes,_emptyData,_dtype)

        _valuePos = len(_full.columns)

        for _row in _full.itertuples():
            _arr=[]
            _isOK = True
            _value = _row[_valuePos]
            for nn in range(1,len(_allIndexNames)+1):
                _indexValue = _row[nn]
                if _indexValue in _allindexes[nn-1]._indices:
                    _pos = _allindexes[nn-1]._indices[_indexValue]
                    _arr.append(_pos)
                else:
                    _isOK=False
                    break

            if _isOK:    
                _cube._values[tuple(_arr)] = _value

        if (not valueIndex is None) and (not replaceByIndex is None):
            _cube = changeIndex(_cube, valueIndex,replaceByIndex, 2 )    

    return _cube 


def indexFromPandas(dataframe, columnName=None, removeEmpty=True, indexName="new index"):
    """ Return a cp.index from an column of a pandas dataframe.
    dataframe: pandas dataframe
    columnName: dataframe column name used for create cp.index. By default is created using the first column
    removeEmpty: True for remove empty rows
        Ex.
            cp.indexFromPandas(df)
            cp.indexFromPandas(df,"column10")
    """                
    
    _serie= None
    if columnName is None:
        _serie=dataframe[dataframe.columns[0]]
    else:
        _serie=dataframe[columnName]

    if removeEmpty:
        _serie.dropna(inplace=True)
        if kindToString( _serie.dtype.kind )=="string" or kindToString( _serie.dtype.kind )=="object":
            _serie = _serie[_serie!=""]

    return index(indexName,_serie.unique())


def pandasFromCube(cube, dataColumnName="value", keepIndexOrder=False ):
    """
    Return indexed Dataframe created from cubePy Cube
    """
    import pandas as pd

    multiIndexValues = [idx.values for idx in cube.axes]
    allIndexNames = [idx.name for idx in cube.axes]
    multiindex = pd.MultiIndex.from_product(multiIndexValues, names=allIndexNames)
    serie = pd.Series(data=cube.values.flatten() , index=multiindex)
    df = pd.DataFrame(serie)

    if keepIndexOrder:
        df = df.reset_index()
        for cubeIndex in cube.axes:
            df[cubeIndex.name] = df[cubeIndex.name].astype(pd.api.types.CategoricalDtype(cubeIndex.values))
        df.set_index(allIndexNames, inplace=True)
    df.columns= [dataColumnName]
    return df


def excel(filepath, useOpenpyxl=False, dataOnly=True, readOnly=True ):
    """ Create excel object from filepath.
    filepath: path to excel file
    useOpenpyxl: True for use custom 
    dataOnly: True for view only the values, not formula
    readOnly: True for read only, False for write options
    Ex.
            cp.excel("\path\to\the\excelfile.xlsx")
    """
    if isLinux():
        filepath = filepath.replace("\\","/")

    if os.path.isfile(filepath):
        if useOpenpyxl: # or isLinux():
            from openpyxl import load_workbook
            return load_workbook(filepath, data_only=dataOnly, read_only=readOnly)
        else:
            return filepath
    else:
        raise ValueError("File not found")    


def pandasFromExcel(excel,sheetName=None,namedRange=None,cellRange=None, indexes=None, driver='Driver={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};DBQ=%s;READONLY=TRUE'):
    """ return a pandas dataframe from excel.
    excel: path to excel file or cp.excel object
    sheetName: sheet name to be read
    namedRange: name of the range to be read
    cellRange: used with sheetname, for read from a simple range
    indexes: Listo of columns names for convert to index of dataframe
        Ex.
            cp.pandasFromExcel(excelNode,"Sheet 1")
            cp.pandasFromExcel(excelNode,namedRange="name_range")
            cp.pandasFromExcel(excelNode,"Sheet 1",cellRange="A1:H10")

    """
    import pandas as pd 
    
    if isinstance(excel,str):

        if isLinux():
            filename=excel
            target_dir = os.path.dirname(filename)
            file_name, file_extension = os.path.splitext(filename)
            target_dir = os.path.join(target_dir, file_name)

            file_to_read = os.path.join(target_dir, namedRange+".pkl")
            if os.path.isfile(file_to_read):
                df = pd.read_pickle(file_to_read,compression='gzip')
                if not indexes is None:
                    df.set_index(indexes,inplace=True)
                return df
            else:
                #TODO: use openpyxl
                raise ValueError(f"Optimized version not found for: '{filename}'")

        else:
            import pyodbc

            cnxn = pyodbc.connect(driver % excel, autocommit=True)
            cursor = cnxn.cursor()
            table = ""
            if not sheetName is None: 
                table = "["+sheetName+"$]"
            if not namedRange is None:
                table = "["+namedRange+"]"
            if not cellRange is None:
                table = "["+sheetName+"$"+cellRange+"]"

            cursor.execute("SELECT * FROM " + table)
            rows = cursor.fetchall()
            columnNames = [str(x[0]) for x in cursor.description]
            cnxn.close()

            df = pd.DataFrame.from_records(rows,columns=columnNames)
            df = df.dropna(how ="all")
            if not indexes is None:
                df.set_index(indexes,inplace=True)
            return df
    else:
        
        if "openpyxl.workbook" in str(type(excel)):
            rangeToRead = None
            if not namedRange is None:
                the_range = excel.defined_names[namedRange]
                dests = the_range.destinations
                for title, coord in dests:
                    ws = excel[title]
                    rangeToRead=ws[coord]
            elif not cellRange is None:
                ws = excel[sheetName]
                rangeToRead = ws[cellRange]
            else: 
                rangeToRead = excel[sheetName]

            nn=0
            cols=[]
            values = []
            for row in rangeToRead:
                if nn==0:
                    cols = [str(c.value) for c in row]
                else:
                    values.append([c.value for c in row]) 
                nn+=1
            nn=0
            _finalCols=[]
            for _col in cols:
                if _col is None:
                    _finalCols.append("Unnamed" + str(nn))
                    nn+=1
                else:
                    _finalCols.append(_col)
            
            df = pd.DataFrame(values,columns=_finalCols)
            if not indexes is None:
                if isinstance(indexes,str):
                    indexes=[indexes]
                toIndex = []
                for indexColumn in indexes:
                    if indexColumn in df.columns.values:
                        toIndex.append(indexColumn)
                if len(toIndex)>0:
                    df.set_index(toIndex, inplace=True)

            return df.dropna(how ="all")
        else:
            raise ValueError("excel can be cp.excel object")        


def indexFromExcel(excel, sheetName=None,namedRange=None,cellRange=None, columnName=None, removeEmpty=True, indexName="new index"):
    """ Return a cp.index from an excel file.
    excel: cp.excel object
    sheetName: sheet name to be read
    namedRange: name of the range to be read
    cellRange: used with sheetname, for read from a simple range
    columnName: dataframe column name used for create cp.index. By default is created using the first column
    removeEmpty: True for remove empty rows
    indexName: new index name 
        Ex.
            cp.indexFromExcel(excelNode,"Sheet 1")
            cp.indexFromExcel(excelNode,namedRange="name_range")
            cp.indexFromExcel(excelNode,namedRange="name_range", columnName="indicadores")
    """                
    if isinstance(excel,str) or "openpyxl.workbook" in str(type(excel)):
        _df = pandasFromExcel(excel,sheetName,namedRange,cellRange)
        return indexFromPandas(_df,columnName,removeEmpty,indexName=indexName)
    else:
        raise ValueError("excel can be cp.excel object or a str path to the filename")


def cubeFromExcel(excel, sheetName=None,namedRange=None,cellRange=None, cubeIndexes=None, valueColumns=None, indexColumnHeaders=None, replaceByIndex=None):
    """ Return a cp.cube from excel file.
        excel: cp.excel object
    sheetName: sheet name to be read
    namedRange: name of the range to be read
    cellRange: used with sheetname, for read from a simple range
    cubeIndexes: objects cp.index for perform change index
    valueColumns: string with column name of the dataframe where contain the values
                    cp.index with columns names for convert colums to index
    indexColumnHeaders: (optional) column names in pandas to parse with cubeIndexes. Used if header on dataframe is not equal to index identifiers.
    replaceByIndex: (optional) replace index used in valueColumns for this index. (using changeindex)

        Ex.
            cp.cubeFromExcel(excelNode,"Sheet 1",cubeIndexes=[indicadores],valueColumns="descuentos")
            cp.cubeFromExcel(excelNode,namedRange="nombre_rango",cubeIndexes=[indicadores],valueColumns=time)
    """

    if isinstance(excel,str) or "openpyxl.workbook" in str(type(excel)):
        _df = pandasFromExcel(excel,sheetName,namedRange,cellRange)
        return cubeFromPandas(_df, cubeIndexes,  valueColumns, indexColumnHeaders, replaceByIndex=replaceByIndex) 
    else:
        raise ValueError("excel can be cp.excel object or a str path to the filename")


def sequence(initialNum, finalNum, stepSize = 1, dtype = None, indexName="new index"): 
    """
    Creates an index with a list of numbers increasing or decreasing from initial_num to final_num by increments (or decrements) of stepSize, which is optional and defaults to 1
    """
    values = np.arange(initialNum, finalNum + 1, step = stepSize, dtype = dtype) 
    return index(indexName,values)


def lookup(cubeWithData, cubeWithMap, sharedIndex): 
    """
    Returns the value of cubeWithData indexed by the index of cubeWithMap.
    cubeWithData must be indexed by sharedIndex and cubeWithData values must correspond to elements of sharedIndex.
    For example: Let's say you have a cube with an estimated inflation rate by Country ("inflation_rate" is the name of the cube; "country" is the name of the index) and you want to assign it to the corresponding Company depending on its location. On the other hand, there's a many-to-one map where each Company is allocated to a single Country ("country_to_company_allocation"). The sharedIndex, in this case, is Country ("country").
    As a result, 
        cp.lookup( inflation_rate , country_to_company_allocation , country )
    will return the estimated inflation rate by Company.
    """
    _final_cube = ((cubeWithMap == sharedIndex) * cubeWithData).sum(sharedIndex) 
    return _final_cube 



def cubeFromNumpy(npArray):
    """ 
    Return a cube object from numpy Array. Generate temporal indexes
    """
    _dimsNames = ["axis " + str(x) for x in range( npArray.ndim)]
    _dimsValues = [ list(x) for x in (range( npArray.shape[y]) for y in range( npArray.ndim)) ]
    _indexes = [cubepy.Index( _dimsNames[x], _dimsValues[x]) for x in range(len(_dimsNames))]
    _cube = cube(_indexes,npArray)
    return _cube


def movingStats(cube, index, window=12, fn=np.mean):
    """
        Allows to calculate in a moving Window the selected statistics measure
    """
    _realWindow = abs(window)
    _fInvert = lambda x: x[::-1].rolling(_realWindow).apply(fn)[::-1].shift(-1)
    _fNorm = lambda x: x.rolling(_realWindow).apply(fn).shift()
    _realF = _fNorm if window<0 else _fInvert
    _levelList = [lvl.name for lvl in cube.axes if lvl.name!=index.name]
    _df = pandasFromCube(cube)
    if cube.ndim>1:
        _df = _df.groupby(level=_levelList)["value"].apply(_realF)
    else:
        _df = _df.apply(_realF)
    return cubeFromPandas( _df, cube.axes, "value")


def isLinux():
    if platform == "linux" or platform == "linux2" or platform == "darwin":
        return True
    else:
        return False
