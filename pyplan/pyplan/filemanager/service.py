import base64
import csv
import json
import os
import shutil
import tempfile
from datetime import datetime
from errno import ENOTDIR
from itertools import islice
from shlex import split
from subprocess import PIPE, Popen
from uuid import uuid4

import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from rest_framework import exceptions

from pyplan.pyplan.common.baseService import BaseService
from pyplan.pyplan.common.utils import _linuxCopy, _unzipFile, _zipFiles
from pyplan.pyplan.companies.models import Company
from pyplan.pyplan.preference.models import Preference
from pyplan.pyplan.user_company_preference.models import UserCompanyPreference

from .classes.fileEntry import FileEntry, eFileTypes
from .classes.fileEntryData import (FileEntryData, eSpecialFileType,
                                    eSpecialFolder)


class FileManagerService(BaseService):

    optimizable_templates = [".xls", ".xlsx", ".xlsm", ".xlsb"]

    def getMainFolders(self):
        """
        list:
        Return a list of all folders.
        """

        result = list()
        company_code = self.client_session.company_code
        # User Workspace
        result.append(
            FileEntry(
                text="My Workspace",
                type=eFileTypes.NOTHING,
                data=FileEntryData(
                    fullPath="",
                    specialFolderType=eSpecialFolder.MY_FOLDER
                )
            )
        )

        # if self.current_user.has_perm("pyplan.view_model_path_root"):
        #     result.append(
        #         FileEntry(
        #             text="Root",
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath="",
        #                 specialFolderType=eSpecialFolder.MODELS_PATH
        #             )
        #         )
        #     )

        # if self.current_user.has_perm("pyplan.view_company_root"):
        #     result.append(
        #         FileEntry(
        #             text=self.client_session.companyName,
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath=company_code,
        #                 specialFolderType=eSpecialFolder.MODELS_PATH
        #             )
        #         )
        #     )
        # else:
        #     # Special Folders
        #     result.append(
        #         FileEntry(
        #             text="Public",
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath=f"{company_code}/Public",
        #                 specialFolderType=eSpecialFolder.PUBLIC
        #             )
        #         )
        #     )

        return result

    def getFoldersAndFiles(self, folder=''):
        """
        list:
        Return a list of all folders.
        """
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        result = list()

        base_path = ""
        if folder.startswith("/"):
            base_path = folder[1:]
        elif folder:
            base_path = f"{folder}/"

        items = storage.listdir(base_path)

        # folders
        for item in sorted(items[0], key=str.lower):
            full_path = os.path.join(base_path, item)
            result.append(
                FileEntry(
                    show=not item.startswith('.'),
                    text=item,
                    type=eFileTypes.MY_FOLDER,
                    data=FileEntryData(
                        fileSize=None,
                        fullPath=full_path,
                        # specialFolderType=eSpecialFolder.MODELS_PATH
                        lastUpdateTime=storage.get_modified_time(
                            full_path),
                    )
                )
            )
        # files
        for item in sorted(items[1], key=str.lower):
            full_path = os.path.join(base_path, item)
            specialFileType = eSpecialFileType.FILE
            lowerItem = item.lower()
            if lowerItem.endswith('.ppl') | lowerItem.endswith('.cbpy') | \
                    lowerItem.endswith('.model') | lowerItem.endswith('.ana'):
                specialFileType = eSpecialFileType.MODEL
            elif lowerItem.endswith('.zip'):
                specialFileType = eSpecialFileType.ZIP

            result.append(
                FileEntry(
                    text=item,
                    type=eFileTypes.PUBLIC,
                    data=FileEntryData(
                        fileSize=storage.size(full_path),
                        fullPath=full_path,
                        extension=full_path[full_path.rfind('.')+1:],
                        specialFileType=specialFileType,
                        lastUpdateTime=storage.get_modified_time(full_path),
                    )
                )
            )
        return result

    def createFolder(self, folder_path, folder_name):
        """
        create:
        Creates a folder inside provided path.
        """
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        full_path = os.path.join(
            storage.base_location, folder_path, folder_name)

        if storage.exists(full_path):
            raise exceptions.NotAcceptable('Folder already exists')
        else:
            os.mkdir(full_path)
            return os.path.join(folder_path, folder_name)

    def createFile(self, my_file, folder_path, name, chunk):
        file_path = os.path.join(
            settings.MEDIA_ROOT, 'models', folder_path, name)
        # Moves file if it already exists
        if chunk is 0 and os.path.isfile(file_path):
            new_name = f"{name}-{datetime.today().strftime('%Y%m%d-%H%M%S')}.old"
            self._copy(
                file_path,
                os.path.join(settings.MEDIA_ROOT, 'models',
                             folder_path, new_name)
            )
            os.remove(file_path)
        # Appends all chunks of this request (chunks of chunks)
        # UI sends multiple requests with multiple chunks each per file
        with open(file_path, 'ab+') as temp_file:
            for chunk in my_file.chunks():
                temp_file.write(chunk)

    def copyFileOrFolder(self, source, destination):
        """
        create:
        Duplicate file or Folder.
        """
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        src = os.path.join(storage.base_location, source)
        dest = os.path.join(storage.base_location, destination)
        if self.isLinux():
            return _linuxCopy(src, dest)
        return self._copy(src, dest)

    def ensureUserWorkspace(self):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        # User Workspace
        if not storage.exists(storage.base_location):
            os.makedirs(storage.base_location)

    def renameFile(self, source, new_name):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        src = os.path.join(storage.base_location, os.path.normpath(source))
        *src_path, src_name = src.rsplit(os.path.sep, 1)
        src_path = ''.join(src_path)
        dest = os.path.join(src_path, new_name)
        os.rename(src, dest)
        return dest

    def duplicateFiles(self, sources):
        result = []
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            source = os.path.normpath(source)
            src = os.path.join(storage.base_location, source)
            *dest_path, dest_name = source.rsplit(os.path.sep, 1)
            dest_path = ''.join(dest_path)
            dest = os.path.join(storage.base_location,
                                dest_path, f'Copy 1 of {dest_name}')
            n = 1
            while storage.exists(dest):
                n += 1
                dest = os.path.join(storage.base_location,
                                    dest_path, f'Copy {n} of {dest_name}')
            if self.isLinux():
                result.append(_linuxCopy(src, dest))
            else:
                result.append(self._copy(src, dest))
        return result

    def moveFiles(self, sources, target):
        result = []
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            source = os.path.normpath(source)
            src = os.path.join(storage.base_location, source)
            *dest_path, dest_name = source.rsplit(os.path.sep, 1)
            dest_path = ''.join(dest_path)
            dest = os.path.join(storage.base_location, target, dest_name)
            result.append(self.recursive_overwrite(src, dest))
            if os.path.isdir(src):
                shutil.rmtree(src)
            else:
                storage.delete(src)
        return result

    def copyFiles(self, sources, target):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            source = os.path.normpath(source)
            src = os.path.join(storage.base_location, source)
            *_, src_name = source.rsplit(os.path.sep, 1)
            *dest_path, dest_name = target.rsplit(os.path.sep, 1)
            dest_path = ''.join(dest_path)
            dest = os.path.join(storage.base_location,
                                dest_path, dest_name, src_name)
            if self.isLinux():
                _linuxCopy(src, dest)
            else:
                self._copy(src, dest)
        return True

    def copyToMyWorkspace(self, source):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        target = os.path.join(
            storage.base_location, self.client_session.company_code, self.current_user.username)
        src = os.path.join(storage.base_location, source)
        if self.isLinux():
            return _linuxCopy(src, target)
        return self._copy(src, target)

    def deleteFiles(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        files = []
        for source in sources:
            full_path = os.path.join(storage.base_location, source)
            if not storage.exists(full_path):
                raise exceptions.NotAcceptable(f'File {source} does not exist')
            else:
                files.append(full_path)
        for file_to_delete in files:
            if os.path.isfile(file_to_delete):
                storage.delete(file_to_delete)
            else:
                shutil.rmtree(file_to_delete)

    def download(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        src_0 = os.path.join(storage.base_location, sources[0])
        if len(sources) is 1 and os.path.isfile(src_0):
            return open(src_0, 'rb'), os.path.relpath(src_0, os.path.join(src_0, '..'))
        else:
            temp = tempfile.SpooledTemporaryFile()
            _zipFiles(sources, storage.base_location, temp,
                      False, None)
            return temp, f"{os.path.relpath(sources[0], os.path.join(sources[0], '..'))}.zip"

    def makeJsonStream(self, json_string: str):
        """
        Returns streamed temp file from json string
        """
        temp = tempfile.SpooledTemporaryFile(mode='w+b')
        temp.write(str.encode(json_string))
        temp.seek(0)
        return temp

    def unzipFile(self, source, target_folder):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        src = os.path.join(storage.base_location, source)
        dest = os.path.join(storage.base_location, target_folder)

        _unzipFile(src, dest)

    def zipFiles(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        zip_file = os.path.join(storage.base_location,
                                f'{os.path.normpath(sources[0])}.zip')

        return _zipFiles(sources, storage.base_location, zip_file, False, None)

    def getHome(self):
        company_id = self.getSession().companyId
        model_path = eSpecialFolder.MODELS_PATH
        res = {}

        filepath = os.path.join(settings.MEDIA_ROOT,
                                'models', 'home.json')
        if os.path.isfile(filepath):
            with open(filepath, "r") as json_file:
                try:
                    res = json.load(json_file)
                    if "tabs" in res:
                        for tab in res["tabs"]:
                            if "folders" in tab:
                                for folder in tab["folders"]:
                                    if "items" in folder:
                                        for item in folder["items"]:
                                            if "image" in item:
                                                image_path = os.path.join(
                                                    settings.MEDIA_ROOT, 'models', item["image"])
                                                if os.path.isfile(image_path):
                                                    with open(image_path, "rb") as f_image:
                                                        item["imagesrc"] = str(
                                                            base64.b64encode(f_image.read()), "utf-8")
                except Exception as ex:
                    raise exceptions.NotAcceptable(ex)
        return res

    def optimizeTemplates(self, sources):
        """Generate csv file for each named range in template for future read.
        """

        if sources:
            preference = Preference.objects.filter(
                code="optimize_templates").first()
            if preference:
                user_company_id = self.client_session.userCompanyId
                c_pref = UserCompanyPreference.objects.filter(
                    user_company_id=user_company_id, preference__code="optimize_templates").first()
                if c_pref:
                    preference.definition = c_pref.definition

            if preference.definition["value"]:
                for template in sources:
                    filename, file_extension = os.path.splitext(template)
                    if file_extension in self.optimizable_templates:
                        template_filename = os.path.join(
                            settings.MEDIA_ROOT, 'models', template)
                        self._generate_csv_from_excel(template_filename)

    # Private

    def _generate_csv_from_excel(self, filename):
        """Generate compressed csv from excel file
        """

        file_name, file_extension = os.path.splitext(filename)
        target_dir = os.path.join(os.path.dirname(filename), file_name)

        if not os.path.isdir(target_dir):
            os.mkdir(target_dir)

        wb = load_workbook(filename, data_only=True, read_only=True)
        for item in wb.defined_names.definedName:
            if not item.is_external and item.type == 'RANGE' and item.attr_text and '!$' in item.attr_text:
                target_filename = os.path.join(target_dir, f'{item.name}.pkl')
                if os.path.isfile(target_filename):
                    os.remove(target_filename)

                dests = item.destinations
                for title, coord in dests:
                    if title in wb:
                        ws = wb[title]
                        rangeToRead = ws[coord]
                        if not isinstance(rangeToRead, tuple):
                            rangeToRead = ((rangeToRead,),)

                        nn = 0
                        cols = []
                        values = []
                        for row in rangeToRead:
                            if nn == 0:
                                cols = [str(c.value) for c in row]
                            else:
                                values.append([c.value for c in row])
                            nn += 1
                        nn = 0
                        _finalCols = []
                        for _col in cols:
                            if _col is None:
                                _finalCols.append(f'Unnamed{str(nn)}')
                                nn += 1
                            else:
                                _finalCols.append(_col)
                        df = pd.DataFrame(values, columns=_finalCols).dropna(how='all')
                        df.to_pickle(target_filename, compression='gzip')

    def recursive_overwrite(self, src, dest, ignore=None):
        if os.path.isdir(src):
            if not os.path.isdir(dest):
                os.makedirs(dest)
            files = os.listdir(src)
            if ignore is not None:
                ignored = ignore(src, files)
            else:
                ignored = set()
            for f in files:
                if f not in ignored:
                    self.recursive_overwrite(
                        os.path.join(src, f),
                        os.path.join(dest, f),
                        ignore
                    )
        else:
            shutil.copy(src, dest)
        return dest

    def _copy(self, src, dest):
        try:
            return shutil.copytree(src, dest)
        except OSError as e:
            # If the error was caused because the source wasn't a directory
            if e.errno == ENOTDIR:
                return shutil.copy(src, dest)
            else:
                raise exceptions.NotAcceptable(
                    'Directory not copied. Error: %s' % e)
        except Exception as e:
            raise e
