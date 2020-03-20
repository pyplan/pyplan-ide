import inspect
import numbers
from sys import exc_info, getsizeof, platform

import numpy as np
import numpydoc
import pandas as pd
import xarray as xr


class XHelpers(object):

    random = -123798
    exact = 1
    start_with = 2
    end_with = 3
    contain = 4
    byVal = 1
    byPos = 2

    @property
    def currentNodeId(self):
        return self.node.identifier

    def __init__(self, node):
        self.node = node

    def release(self):
        self.node = None

    def dataArray(self, coords, values=None, dtype=None):
        """DEPRECATED: Use pp.cube"""
        return self.cube(coords, values, dtype)

    def cube(self, coords, values=None, dtype=None):
        """Creat a cube
            TODO: samples
        """
        if not isinstance(coords, list):
            coords = [coords]

        # convert to realCoords ("name",[values])
        _realCoords = None
        if isinstance(coords[0], tuple):
            _realCoords = coords
        elif isinstance(coords[0], XIndex):
            _realCoords = [xx.coord for xx in coords]

        # shape of coords
        _shape = [len(xx[1]) for xx in _realCoords]

        if values is None:
            if not dtype is None:
                if dtype is str:
                    values = np.empty(_shape, dtype='O')
                    values[:] = ""
                    return xr.DataArray(values, _realCoords)
                elif self.kindToString(np.dtype(dtype).kind) == "string":
                    values = np.empty(_shape, dtype=dtype)
                    values[:] = ""
                    return xr.DataArray(values, _realCoords)
                else:
                    values = np.empty(_shape, dtype='float64')
                    values[:] = 0
                    return xr.DataArray(values, _realCoords)
            else:
                values = np.empty(_shape, dtype='float64')
                values[:] = 0
                return xr.DataArray(values, _realCoords)
        else:
            if isinstance(values, list) or isinstance(values, np.ndarray):
                if len(values) > 0:
                    if isinstance(values[0], xr.DataArray):
                        # use stack
                        if isinstance(coords, list):
                            coords = coords[0]
                        return xr.concat(values, dim=coords)

                # TODO: check for correct size
                values = self._checkAndReshape(values, _shape, dtype)
                return xr.DataArray(values, _realCoords)
                # return cubepy.Cube(axes,values,fillValues=True, dtype=dtype)

            elif isinstance(values, numbers.Number) and values == self.random:
                values = np.random.randint(100, size=_shape)
                return xr.DataArray(values, _realCoords)
            else:
                if isinstance(values, np.ndarray) or isinstance(values, list):
                    return xr.DataArray(values, _realCoords)
                else:
                    # recreate data
                    values = self._checkAndReshape(values, _shape, dtype)
                    return xr.DataArray(values, _realCoords)

    def index(self, values, parent=None, mapArrayId=None, dtype=None):
        """Create a coord object from list of values
            TODO: samples        
        """
        return XIndex(self.node.identifier, values, dtype, parent, mapArrayId)

    def createTime(self, date_start, date_end, freq='M', format='%Y.%m'):
        """Create time index usign start and end dates and freq. The result is formated to format parameter
            Ex.
                cp.createTime('2016.01','2018.12')
                cp.createTime('2016.01','2016.12',freq='D',format='%d/%m/%Y')
        """
        if "." in date_start:
            date_start = date_start.replace('.', '-')
        if "." in date_end:
            date_end = date_end.replace('.', '-')
        return self.index(pd.period_range(start=date_start, end=date_end, freq=freq).strftime(format))

    def selfArray(self, index):
        """Create an array from index
        """
        return self.cube([index], index.values)

    def changeIndex(self, array, oldIndex, newIndex, compareMode=1, defaultValue=None):
        """ Change index of a dataArray object.
        """
        _da = array

        if compareMode == 1:
            _temp = _da.reindex({oldIndex.name: newIndex.values})
            _temp[newIndex.name] = _temp[oldIndex.name]
            _temp = _temp.swap_dims(
                {oldIndex.name: newIndex.name}).drop(oldIndex.name)
            if not defaultValue is None:
                _temp = _temp.fillna(defaultValue)

            return _temp

        else:

            if len(oldIndex.values) == len(newIndex.values):
                _tmp = _da.copy()
                _tmp.coords[newIndex.name] = _tmp.coords[oldIndex.name]
                _tmp.coords[newIndex.name].data = newIndex.values
                _tmp = _tmp.swap_dims(
                    {oldIndex.name: newIndex.name}).drop(oldIndex.name)
                return _tmp
            elif len(oldIndex.values) > len(newIndex.values):
                raise ValueError(
                    "Changeindex by pos for indices of different size is not implemented")
            else:
                raise ValueError(
                    "Changeindex by pos for indices of different size is not implemented")

    def cubeFromPandas(self, dataframe, columnName=None, indexMap=None, aggFn=None, defaultValue=0):
        """Create new pp.cube, converting pandas to multidimensional data, according to the parameters
            dataframe: pandas dataframe
            columnName: string with dataframe column name where contain the values
            indexMap: dictionary with dataframe column name as key and Pyplan index as value
                Ex.
                    pp.cubeFromPandas(simple_pandas)
                    pp.cubeFromPandas(simple_pandas,columnName="col1")
                    pp.cubeFromPandas(simple_pandas,columnName="col1", indexMap={"A":index_a})
                    pp.cubeFromPandas(simple_pandas,columnName="col1", indexMap={"A":index_a, "B":index_b}, aggFn=np.max)
                    pp.cubeFromPandas(simple_pandas["col1"])
                    pp.cubeFromPandas(simple_pandas["col2"],indexMap={"A":index_a, "B":index_b})
        """
        if aggFn is None:
            aggFn = np.sum
        res = None
        if isinstance(dataframe, pd.Series):
            dataframe = pd.DataFrame({"value": dataframe})
            columnName = "value"

        if isinstance(dataframe, pd.DataFrame):
            if not columnName:
                columnName = dataframe.columns[0]

            if indexMap is None:
                res = dataframe[columnName].to_xarray()
            else:
                df = dataframe.reset_index()
                newIndexes = []
                for key in indexMap:
                    indexMap[key] = indexMap[key].name
                    newIndexes.append(indexMap[key])
                keys = [xx for xx in indexMap]
                allColums = list(keys)
                allColums.append(columnName)
                df = df[allColums]
                df.rename(columns=indexMap, inplace=True)
                df.set_index(newIndexes, inplace=True)
                df = df.groupby(newIndexes).agg(aggFn)
                res = df[columnName].to_xarray()
                df = None

        res = res.fillna(defaultValue)

        return res

    def indexFromPandas(self, dataframe, columnName=None, removeEmpty=True):
        """ Return a pp.index from an column of a pandas dataframe.
        dataframe: pandas dataframe
        columnName: dataframe column name used for create cp.index. By default is created using the first column
        removeEmpty: True for remove empty rows
            Ex.
                cp.indexFromPandas(df)
                cp.indexFromPandas(df,"column10")
        """

        _serie = None
        if columnName is None:
            _serie = dataframe[dataframe.columns[0]]
        else:
            _serie = dataframe[columnName]

        if removeEmpty:
            _serie.dropna(inplace=True)
            if self.kindToString(_serie.dtype.kind) == "string" or self.kindToString(_serie.dtype.kind) == "object":
                _serie = _serie[_serie != ""]

        return self.index(_serie.unique())

    def excelConnection(self, filepath, useOpenpyxl=False, dataOnly=True, readOnly=True):
        """ Create excel object from filepath.
        filepath: path to excel file
        useOpenpyxl: True for use custom 
        dataOnly: True for view only the values, not formula
        readOnly: True for read only, False for write options
        Ex.
                pp.excel("\path\to\the\excelfile.xlsx")
        """
        from pathlib import Path

        fullFilename = filepath
        my_file = Path(fullFilename)
        if not my_file.is_file():
            fullFilename = self.node.model.getNode(
                "current_path").result + filepath

        my_file = Path(fullFilename)
        if my_file.is_file():
            if useOpenpyxl or self.isLinux():
                from openpyxl import load_workbook
                return load_workbook(fullFilename, data_only=dataOnly, read_only=readOnly)
            else:
                return filepath
        else:
            raise ValueError("File not found")

    def pandasFromExcel(self, excel, sheetName=None, namedRange=None, cellRange=None, indexes=None, driver='Driver={Microsoft Excel Driver (*.xls, *.xlsx, *.xlsm, *.xlsb)};DBQ=%s;READONLY=TRUE'):
        """ return a pandas dataframe from excel.
        excel: path to excel file or cp.excel object
        sheetName: sheet name to be read
        namedRange: name of the range to be read
        cellRange: used with sheetname, for read from a simple range
        indexes: Listo of columns names for convert to index of dataframe
            Ex.
                cp.pandasFromExcel(excelNode,"Sheet 1")
                cp.pandasFromExcel(excelNode,namedRange="name_range")
                cp.pandasFromExcel(excelNode,"Sheet 1",cellRange="A1:H10")

        """
        import pandas as pd

        if isinstance(excel, str):
            from pathlib import Path
            my_file = Path(excel)
            if not my_file.is_file():
                excel = self.node.model.getNode("current_path").result + excel

            import pyodbc

            cnxn = pyodbc.connect(driver % excel, autocommit=True)
            cursor = cnxn.cursor()
            table = ""
            if not sheetName is None:
                table = "["+sheetName+"$]"
            if not namedRange is None:
                table = "["+namedRange+"]"
            if not cellRange is None:
                table = "["+sheetName+"$"+cellRange+"]"

            cursor.execute("SELECT * FROM " + table)
            rows = cursor.fetchall()
            columnNames = [str(x[0]) for x in cursor.description]
            cnxn.close()

            df = pd.DataFrame.from_records(rows, columns=columnNames)
            df = df.dropna(how="all")
            if not indexes is None:
                df.set_index(indexes, inplace=True)
            return df
        else:

            if "openpyxl.workbook" in str(type(excel)):
                rangeToRead = None
                if not namedRange is None:
                    the_range = excel.defined_names[namedRange]
                    dests = the_range.destinations
                    for title, coord in dests:
                        ws = excel[title]
                        rangeToRead = ws[coord]
                elif not cellRange is None:
                    ws = excel[sheetName]
                    rangeToRead = ws[cellRange]
                else:
                    rangeToRead = excel[sheetName]

                nn = 0
                cols = []
                values = []
                for row in rangeToRead:
                    if nn == 0:
                        cols = [c.value for c in row]
                    else:
                        values.append([c.value for c in row])
                    nn += 1
                df = pd.DataFrame(values, columns=cols)
                if not indexes is None:
                    if isinstance(indexes, str):
                        indexes = [indexes]
                    toIndex = []
                    for indexColumn in indexes:
                        if indexColumn in df.columns.values:
                            toIndex.append(indexColumn)
                    if len(toIndex) > 0:
                        df.set_index(toIndex, inplace=True)

                return df.dropna(how="all")
            else:
                raise ValueError("excel can be cp.excel object")

    def indexFromExcel(self, excel, sheetName=None, namedRange=None, cellRange=None, columnName=None, removeEmpty=True):
        """ Return a pp.index from an excel file.
        excel: pp.excel object
        sheetName: sheet name to be read
        namedRange: name of the range to be read
        cellRange: used with sheetname, for read from a simple range
        columnName: dataframe column name used for create pp.index. By default is created using the first column
        removeEmpty: True for remove empty rows
            Ex.
                pp.indexFromExcel(excelNode,"Sheet 1")
                pp.indexFromExcel(excelNode,namedRange="name_range")
                pp.indexFromExcel(excelNode,namedRange="name_range", columnName="indicadores")
        """
        if isinstance(excel, str) or "openpyxl.workbook" in str(type(excel)):
            _df = self.pandasFromExcel(excel, sheetName, namedRange, cellRange)
            return self.indexFromPandas(_df, columnName, removeEmpty)
        else:
            raise ValueError(
                "excel can be pp.excel object or a str path to the filename")

    def dataArrayFromExcel(self, excel, sheetName=None, namedRange=None, cellRange=None, indexes=None, valueColumns=None, indexColumnHeaders=None, replaceByIndex=None):
        """ DEPRECATED: Use pp.cubeFromExcel instead.
        """
        return self.cubeFromExcel(excel, sheetName, namedRange, cellRange, indexes, valueColumns, indexColumnHeaders, replaceByIndex)

    def cubeFromExcel(self, excel, sheetName=None, namedRange=None, cellRange=None, indexes=None, valueColumns=None, indexColumnHeaders=None, replaceByIndex=None):
        """ Return a pp.cube from excel file.
            excel: pp.excel object
        sheetName: sheet name to be read
        namedRange: name of the range to be read
        cellRange: used with sheetname, for read from a simple range
        indexes: objects pp.index for perform change index
        valueColumns: string with column name of the dataframe where contain the values
                        pp.index with columns names for convert colums to index
        indexColumnHeaders: (optional) column names in pandas to parse with indexes. Used if header on dataframe is not equal to index identifiers.
        replaceByIndex: (optional) replace index used in valueColumns for this index. (using changeindex)

            Ex.
                pp.dataArrayFromExcel(excelNode,"Sheet 1",indexes=[indicadores],valueColumns="descuentos")
                pp.dataArrayFromExcel(excelNode,namedRange="nombre_rango",indexes=[indicadores],valueColumns=time)
        """

        dataframe = self.pandasFromExcel(
            excel, sheetName, namedRange, cellRange)
        import pandas as pd

        valueIndex = None
        if isinstance(valueColumns, XIndex):
            valueIndex = valueColumns
            valueColumns = valueIndex.values
        elif isinstance(valueColumns, str):
            valueColumns = [valueColumns]

        if indexColumnHeaders is None:
            indexColumnHeaders = [index.name for index in indexes]

        # create total index and index names

        _allindexes = indexes
        _allIndexNames = indexColumnHeaders[:]
        if not valueIndex is None:
            _allindexes.append(valueIndex)
            _allIndexNames.append("data_index")

        _full = dataframe.reset_index().melt(id_vars=indexColumnHeaders,
                                             value_vars=valueColumns, var_name="data_index", value_name="data_value")

        # sum for acum over duplicate records
        _full = _full.groupby(_allIndexNames, as_index=False).sum()
        _dtype = _full["data_value"].dtype

        _dataType = self.kindToString(_dtype.kind)
        if _dataType == "string":
            _full = _full[(_full["data_value"] != "") &
                          (_full['data_value'].notna())]
        else:
            _full = _full[(_full["data_value"] != 0) &
                          (_full['data_value'].notna())]

        _full.set_index(_allIndexNames, inplace=True)
        _da = _full["data_value"].to_xarray()

        # if indexed, rename index
        if not indexes is None and not indexColumnHeaders is None:
            if not isinstance(indexes, list):
                indexes = [indexes]
            idxPos = 0
            for cubeIndex in indexes:
                newIndexName = cubeIndex.name
                if idxPos <= len(indexColumnHeaders)-1:
                    oldIndexName = indexColumnHeaders[idxPos]
                    if not newIndexName in _da.coords:
                        _da.coords[newIndexName] = _da.coords[oldIndexName]
                        _da = _da.swap_dims(
                            {oldIndexName: newIndexName}).drop(oldIndexName)
                    idxPos += 1
                    # reindex for create complete combinations
                    _da = _da.reindex({newIndexName: cubeIndex.values})

        if not valueIndex is None:
            newIndexName = valueIndex.name
            oldIndexName = "data_index"
            if not newIndexName in _da.coords:
                _da.coords[newIndexName] = _da.coords[oldIndexName]
                _da = _da.swap_dims(
                    {oldIndexName: newIndexName}).drop(oldIndexName)
            # reindex for create complete combinations
            _da = _da.reindex({newIndexName: valueIndex.values})

            if not replaceByIndex is None:
                _da = self.changeIndex(_da, valueIndex, replaceByIndex, 2)

        return _da

    def concatIndex(self, *args):
        """
        Concatenates two or more indexes and/or atomic values into a single new index
        Return: new index
            Ex.
                cp.concatIndex(index1,index2,index3,value1,value2)
        """
        _list = []
        for arg in args:
            if isinstance(arg, XIndex):
                values = (arg.values).tolist()
                _list.extend(values)
            else:
                _list.append(arg)

        seripandas = pd.Series(_list)
        return self.index(seripandas.unique())

    # ahora es una funcion en la libreria externa
    # def yearsToTime(self,dataArray,yearsIndex,timeIndex,div=12):
    #     """ Convert dataArray indexed by anual index to a monthly index
    #     dataArray are the dataArray with annual values to be converted
    #     yearIndex: is the year index of the original matrix
    #     timeIndex: is the monthly index of the result
    #     div: is an optional parameter used as divisor of the annual values (the tipical value is 12)
    #     """
    #     matrix = self.find( yearsIndex, timeIndex, compareType = 4 )
    #     final_cube = (dataArray * matrix).sum(yearsIndex.name)  / div
    #     return final_cube

    def find(self, param1, param2, compareType=1, caseSensitive=True):
        """
        param1: value or indexarray for compare
        param2: index compare to
        compareType: cp.exact=1, cp.start_with=2, cp.end_with=3, cp.contain=4  
        caseSensitive: able to differentiate between uppercase and lowercase (by default True)

        If param1 is a scalar (numeric or str) and param2 is an index:  return a dataArray indexed by param2 with True on ocurrences of param2
            Ex. result = pp.find("te", region, cp.end_with)
        If param1 is an index and param2 is an index too:  return a dataArray indexed by param1 and param2 with True on ocurrences of param1 on param2
            Ex. result = pp.find(subregion, region, cp.contain)

        """
        def _fn(item, value):
            if isinstance(item, str) == False:
                item = str(item)
            if isinstance(value, str) == False:
                value = str(value)
            if compareType == 1:
                if caseSensitive:
                    return item == value
                else:
                    return item.lower() == value.lower()
            elif compareType == 2:
                if caseSensitive:
                    return item[:len(value)] == value
                else:
                    return item[:len(value)].lower() == value.lower()
            elif compareType == 3:
                if caseSensitive:
                    return item[-len(value):] == value
                else:
                    return item[-len(value):].lower() == value.lower()
            elif compareType == 4:
                if caseSensitive:
                    return value in item
                else:
                    return value.lower() in item.lower()

        if (isinstance(param1, str) or str(param1).isnumeric()) and isinstance(param2, XIndex):
            vfn = np.vectorize(_fn)
            return self.cube([param2], vfn(param2.values, param1))

        if isinstance(param1, XIndex) and isinstance(param2, XIndex):
            _res = self.cube([param1, param2], False, dtype=bool)
            for row in param1.values:
                for col in param2.values:
                    _res.loc[{param1.name: slice(row, row), param2.name: slice(col, col)}] = _fn(
                        col, row)
            return _res

    def splitText(self, param1, separator, part=None):
        """
        Description: Returns a DataArray object with text values formed by splitting the elements of param1 text values at each occurrence of separator "separator". 
        The DataArray will have the original dimension plus a new dimension 'Parts' of length (number of separators + 1). 
        All text values must have the same number of separators separator.        
        """
        if isinstance(param1, XIndex):
            param1 = self.cube(param1, param1.values)

        _q_separators = self.apply(param1, lambda x: x.count(separator))
        _max_q_separators = np.asscalar(_q_separators.max().values)
        _result_coords = ['Part ' + str(i)
                          for i in range(1, _max_q_separators + 2)]
        _result_dim = self.index(_result_coords)
        _result_dim.name = "Parts"

        _results = []

        for _part in range(_max_q_separators + 1):
            _dataarray = self.apply(
                param1, lambda x: x.split(separator)[_part])
            _results.append(_dataarray)

        _res = xr.concat(_results, dim=_result_dim)

        if not part is None:
            _res = _res.sel(Parts="Part " + str(part), drop=True)

        return _res

    def choice(self, index, selection, includeAll=False):
        """Return the element in the "selection" position of the index. 
        """
        if selection == 0 and includeAll == 1:
            return "All"
        else:
            values = None
            if isinstance(index, XIndex):
                values = (index.values[:1000]).tolist()
            elif isinstance(index, np.ndarray):
                values = (index[:1000]).tolist()
            else:
                values = list(index)[:1000]
            if not values is None and len(values) >= selection:
                return values[selection-1]
        return ""

    def copyAsValues(self, source, targetId):
        """Copy values of datArray "source" into dataArray with id 'targetId'. This function alter the definition of dataArray with 'targetId' identifier.
            source: dataArray/index from copy values
            targetId: identifier (string) of the target node 
        """

        if isinstance(source, str):
            if self.node.model.existNode(source):
                source = self.node.model.getNode(source).result
            else:
                raise ValueError("Node '" + source + "' not found")

        if not isinstance(source, xr.DataArray) and not isinstance(source, XIndex) and not isinstance(source, float) and not isinstance(source, int):
            raise ValueError(
                "The 'source' parameter must be a DataArray, Index, float or int")

        if not isinstance(targetId, str):
            raise ValueError(
                "The 'targetId' parameter must be a string (identifier of node)")

        if not self.node.model.existNode(targetId):
            raise ValueError("Node '" + targetId + "' not found")

        newDef = ""
        if isinstance(source, float) or isinstance(source, int):
            newDef = "result = " + str(source)
        elif isinstance(source, xr.DataArray):
            indexes = str(list(source.dims)).replace("'", '')
            np.set_printoptions(threshold=np.prod(source.values.shape))
            data = np.array2string(source.values, separator=",", precision=20, formatter={
                'float_kind': lambda x: "np.nan" if np.isnan(x) else repr(x)}).replace('\n', '')
            newDef = "result = pp.dataArray(" + indexes + "," + data + ")"
        elif isinstance(source, XIndex):
            np.set_printoptions(threshold=np.prod(source.values.shape))
            data = np.array2string(source.values, separator=",", precision=20, formatter={
                'float_kind': lambda x: "np.nan" if np.isnan(x) else repr(x)}).replace('\n', '')
            newDef = "result = pp.index(" + data + ")"

        self.node.model.getNode(targetId).definition = newDef
        return True

    def goalSeek(self, nodeIdX, nodeIdObjective, goal=0, startValue=1):
        """ Finds the value of nodeIdX that makes nodeIdObjective equal to goal.
        nodeIdX: String with id of node X
        nodeIdObjective: String with id of node X
        """
        from scipy.optimize import newton

        def _f(x):
            self.node.model.getNode(nodeIdX).definition = "result = " + str(x)
            value = self.node.model.getNode(nodeIdObjective).result
            return value - goal

        _res = newton(_f, x0=startValue)
        return _res

    def plot(self, fig, **kwargs):
        """
        Generate HTML with plotly chart
        Params:
            fig: plotly fig
        """
        import plotly.io as pio
        return pio.to_html(fig, full_html=False, **kwargs)

    def aggregate(self, dataArray, mapInfo, sourceIndex, targetIndex):
        """ Aggregates the values in DataArray to generate the result indexed by  targetIndex.
            Map gives the value of targetIndex for each element of sourceIndex (If the map does not match then the element will not be set into target index and information will be lost)

            Example for aggregating time information into annual index the syntax is:
                cp.aggregate(dataArray, map, time, years )
        """
        _map = pd.DataFrame({targetIndex.name: mapInfo.to_series()})
        _serie = dataArray.to_series()
        _df = pd.DataFrame({"value": _serie})
        _df = _df[_df["value"] != 0]
        _df = _df.join(_map).reset_index()
        _df.drop(columns=[sourceIndex.name], inplace=True)
        _newDimList = [
            xx for xx in dataArray.dims if xx not in [sourceIndex.name]]
        _newDimList.append(targetIndex.name)
        _df = _df.groupby(_newDimList).sum()
        _da = _df["value"].to_xarray()

        # reindex
        _reindexDic = {targetIndex.name: targetIndex.values}
        for coord in dataArray.coords:
            if coord != sourceIndex.name:
                _reindexDic[coord] = dataArray.coords[coord].values
        _da = _da.reindex(_reindexDic)
        # fin reindex
        return _da.fillna(0)

    def nestedGroupby(self, dataArray, groupby, applyFn):
        """ Apply "applyFn" to all dims in "groupby" of the "dataArray"
        """
        if len(groupby) == 1:
            return dataArray.groupby(groupby[0]).apply(applyFn)
        else:
            return dataArray.groupby(groupby[0]).apply(self.nestedGroupby, groupby=groupby[1:], applyFn=applyFn)

    def subset(self, cube):
        """Returns a list of all the elements of the index for which cube is true. The function is used to create a new index that is a subset of an existing index.
            Ex. pp.subset(sales>0)
        """
        cond = cube > 0
        values = cond.coords[cond.dims[0]].values[cond.values]
        return self.index(values)

    def apply(self, obj, applyFn, *args):
        """ Apply "applyFn" to "obj" where obj can be DataArray or Index
        """
        vfn = np.vectorize(applyFn)
        if isinstance(obj, XIndex):
            return self.index(vfn(obj.values, *args))
        if isinstance(obj, xr.DataArray):
            return xr.apply_ufunc(vfn, obj, *args)
        return None

    def sel(self, dataArray, filterList, compareMode=1, defaultValue=None):
        """
        Filter dataArray using the filterList filters. 

        dataArray: dataArray to be filtered
        filterList: the possible filters are:
                index==value
                index1==indice2 (in this case a changeindex will be made).
        compareMode: 1: by Value (default), 2: by pos (used only if changeindex is necessary)
        defaultValue: value to fill the elements that are not found (used only if changeindex is necessary) 
            Ex.
                pp.sel(dataArray1, index==value)
                pp.sel(dataArray1, [index1==value1, index2==value2])
                pp.sel(dataArray1, [index1==value1, index2==index3])
        """
        if not isinstance(dataArray, xr.DataArray):
            raise ValueError(
                "the 'dataArray' parameter must be of the type xr.DataArray")

        if not isinstance(filterList, list):
            filterList = [filterList]

        res = dataArray
        for filterItem in filterList:
            if filterItem.ndim == 1:
                res = res.where(filterItem, drop=True).squeeze(drop=True)
            if filterItem.ndim == 2:
                oldIndex = filterItem.coords[filterItem.dims[0]]
                newIndex = filterItem.coords[filterItem.dims[1]]
                res = self.changeIndex(
                    res, oldIndex, newIndex, compareMode, defaultValue)

        return res

    def isel(self, dataArray, index, position):
        """
        Filter dataArray by integer position along the specified index.

        dataArray: dataArray to be filtered
        index: pp.index 
        position: int 
            Ex.
                pp.isel(dataArray1, index1, 0)
        """
        if not isinstance(dataArray, xr.DataArray):
            raise ValueError(
                "the 'dataArray' parameter must be of the type xr.DataArray")
        return dataArray.isel({index.name: position}, drop=True)

    def irr(self, flow, time_index):
        """Returns the Internal Rate of Return (IRR) of a series of periodic payments (negative values) and inflows (positive values). The IRR is the discount rate at which the Net Present Value (NPV) of the flows equals zero. 
            The variable flow must be indexed by time_index.

        If the cash flow never changes sign, cp.irr() has no solution and returns NAN (Not A Number).
        """
        raise ValueError("No implementado, se buscan valientes")
        import pandas as pd

        _cube_dimensions = index("flowdims", flow.dims)
        _rest_of_indexes_labels = subset(
            _cube_dimensions != time_index.name).values
        _rest_of_indexes = [flow.axis(xx) for xx in _rest_of_indexes_labels]

        _cube = None
        if len(_rest_of_indexes) == 0:
            _cube = np.irr(flow.values)
        else:
            _cube = cube(_rest_of_indexes)
            _multivalues = [idx.values for idx in _rest_of_indexes]

            _values = pd.MultiIndex.from_product(_multivalues).values

            for _item in _values:
                _filter = []

                for _nn in range(len(_item)):
                    _filter.append(_rest_of_indexes[_nn].filter([_item[_nn]]))

                _irr = np.irr(flow.filter(_filter).squeeze().values)
                _cube.set_data(_filter, _irr)
        return _cube

    def npv(self, rate, flow, time_index, offset=1):
        """"Returns the Net Present Value (NPV) of a cash flow with equally spaced periods. The flow parameter must contain a series of periodic payments (negative values) and inflows (positive values), indexed by time_index.
            The optional offset parameter especifies the offset of the first value relative to the current time period. By default, offset is set to 1, indicating that the first value is discounted as if it is one step in the future
        """
        _number_of_periods = time_index.pos + offset
        _present_values = flow / (1 + rate) ** _number_of_periods
        _npv = _present_values.sum(time_index.name)
        return _npv

    def lookup(self, dataArray, dataMap, sharedIndex, defaultValue=0):
        """
        Returns the value of dataArray indexed by the index of dataMap.
        dataArray must be indexed by sharedIndex and dataArray values must correspond to elements of sharedIndex.
        For example: Let's say you have a cube with an estimated inflation rate by Country ("inflation_rate" is the name of the cube; "country" is the name of the index) and you want to assign it to the corresponding Company depending on its location. On the other hand, there's a many-to-one map where each Company is allocated to a single Country ("country_to_company_allocation"). The sharedIndex, in this case, is Country ("country").
        As a result, 
            pp.lookup( inflation_rate , country_to_company_allocation , country )
        will return the estimated inflation rate by Company.
        """

        try:
            return dataArray.sel({sharedIndex.name: dataMap}, drop=True)
        except Exception as ex:
            filtro = dataMap.isin(sharedIndex.values)
            valuesOk = dataMap[dataMap.isin(sharedIndex.values)]
            lookOk = dataArray.sel({sharedIndex.name: valuesOk}, drop=True)
            final = lookOk.reindex(
                {dataMap.dims[0]: dataMap.coords[dataMap.dims[0]].values})
            return self.fillna(final, defaultValue)

    def copyIndex(self, dataArray):
        """Generate a pp.index with current unique values of the dataArray. 
        The dataArray must have only one dimension
        """
        import pandas as pd
        if dataArray.values.ndim > 1:
            raise ValueError("The dataArray must have only one dimension")

        np_values = dataArray.values.flatten()
        seripandas = pd.Series(np_values)
        return self.index(seripandas.unique())

    def dynamic(self, dataArray, index, shift, initialValues=None):
        """
        Perform cyclic calculations betwwen nodes.
        cube: cp.cube to evaluate
        index: Index to shift 
        shift: amount of elemnts to shift. Can be positive or negative
        initialValues: (optional), initial values to apply to first "shift" elemnts
        """
        _da = dataArray.shift({index.name: (shift*-1)})
        if not initialValues is None:
            _da = _da.fillna(initialValues)
        return _da

    def kindToString(self, kind):
        """Returns the data type on human-readable string
        """
        if kind in {'U', 'S'}:
            return "string"
        elif kind in {'b'}:
            return "boolean"
        elif kind in {'i', 'u', 'f', 'c'}:
            return "numeric"
        elif kind in {'m', 'M'}:
            return "date"
        elif kind in {'O'}:
            return "object"
        elif kind in {'V'}:
            return "void"

    def _checkAndReshape(self, values, shape, dtype):
        """
        check if values has coorect size, else return reshaped array
        """
        values = np.asarray(values, dtype)
        valueSize = values.size
        coorSize = 1
        for x in shape:
            coorSize = coorSize * x
        if coorSize != valueSize:
            if coorSize > valueSize:
                if values.dtype.kind in {'U', 'S'}:
                    values = np.append(values.reshape(values.size), [
                                       ''] * (coorSize-values.size)).reshape(shape)
                elif values.dtype.kind in {'b'}:
                    values = np.append(values.reshape(values.size), [
                                       False] * (coorSize-values.size)).reshape(shape)
                else:
                    values = np.append(values.reshape(values.size), np.zeros(
                        coorSize-values.size)).reshape(shape)
            else:
                values = values.flatten()[:coorSize].reshape(shape)
        else:
            # compare shapes
            isok = True
            if len(shape) != len(values.shape):
                isok = False
            if isok:
                for ii in range(len(shape)):
                    if shape[ii] != values.shape[ii]:
                        isok = False
                        break
            if not isok:
                values = values.flatten()[:coorSize].reshape(shape)
        return values

    def from_cpindex(self, cpindex, dtype=None, parent=None, mapArrayId=None):
        """Create a xarray index from cubepy index
        """
        return XIndex(self.node.identifier, cpindex.values, dtype, parent, mapArrayId)

    def from_cpcube(self, coords, cpcube):
        """Create a dataarray from cubepy cube
        """
        return self.cube(coords, values=cpcube.values)

    def addPeriods(self, start, periods, freq='M', format='%Y.%m'):
        """Add periods to a date. Can set freq and output format 
            Ex.
                pp.addPeriods('2016.01',6)
                pp.apply( pp.addPeriods, inicio_de_proyectos , duracin_de_proyectos)
        """
        if "." in start:
            start = start.replace('.', '-')
        if periods < 0:
            return pd.period_range(end=start, periods=-periods+1, freq=freq).strftime(format)[0]
        else:
            return pd.period_range(start=start, periods=periods+1, freq=freq).strftime(format)[-1]

    def minimum(self, dataArray1, dataArray2):
        """Finds the minimum between two DataArrays comparing element per element
            Ex.
                pp.minimum(dataArray1,dataArray2)
        """
        return xr.ufuncs.minimum(dataArray1, dataArray2)

    def maximum(self, dataArray1, dataArray2):
        """Finds the maximum between two DataArrays comparing element per element
            Ex.
                pp.maximum(dataArray1,dataArray2)
        """
        return xr.ufuncs.maximum(dataArray1, dataArray2)

    def fillna(self, dataArray, value=0):
        """TODO: comment
            Ex.
                pp.fillna(dataArray,0)
        """
        return dataArray.fillna(value)

    def fillInf(self, dataArray, value=0):
        """TODO: comment
            Ex.
                pp.fillInf(dataArray,0)
        """
        return self.apply(dataArray, lambda x: value if np.isinf(x) else x)

    def fillAll(self, dataArray, value=0):
        """TODO: comment
            Ex.
                pp.fillAll(dataArray,0)
        """
        return self.fillInf(self.fillna(dataArray, value), value)

    def where(self, cond, x, y):
        """ Return elements from `x` or `y` depending on `cond`.
        Performs xarray-like broadcasting across input arguments.

        cond : scalar, array or DataArray with boolean dtype
                When True, return values from `x`, otherwise returns values from `y`.
        x, y : scalar, array or DataArray
                Values from which to choose. All dimension coordinates on these objects must be aligned with each other and with `cond`.
            Ex. 
                pp.where(dataArray1>10,dataArray1,0)
                pp.where(index1=="item 1",dataArray1,dataArray2)
        """
        return xr.where(cond, x, y)

    def size(self, x):
        """ Return the size of the x element.
        x: Index or dataArray
            Ex:
                pp.size(dataArray1)
                pp.size(index1)
        """
        if isinstance(x, XIndex):
            return x.values.size
        elif isinstance(x, xr.DataArray):
            return x.size
        else:
            raise ValueError("the 'x' parameter must be a DataArray or Index")

    def _getDimsToOperate(self, dataArray, index=None, keep=None):
        dims = None
        if not index is None:
            if not isinstance(index, list):
                index = [index]
            dims = [x.name for x in index]
        elif not keep is None:
            if not isinstance(keep, list):
                keep = [keep]
            dims = []
            keepDims = [x.name for x in keep]
            for dim in dataArray.dims:
                if not dim in keepDims:
                    dims.append(dim)
        return dims

    def sum(self, dataArray, index=None, keep=None):
        """Reduce this DataArray's data by applying 'sum' along some index'es
            Ex:
                pp.sum(dataArray)
                pp.sum(dataArray,index1)
                pp.sum(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.sum(dims)

    def prod(self, dataArray, index=None, keep=None):
        """Reduce this DataArray's data by applying 'prod' along some index'es
            Ex:
                pp.prod(dataArray)
                pp.prod(dataArray,index1)
                pp.prod(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.prod(dims)

    def max(self, dataArray, index=None, keep=None):
        """Reduce this DataArray's data by applying 'max' along some index'es
            Ex:
                pp.max(dataArray)
                pp.max(dataArray,index1)
                pp.max(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.max(dims)

    def min(self, dataArray, index=None, keep=None):
        """Reduce this DataArray's data by applying 'min' along some index'es
            Ex:
                pp.min(dataArray)
                pp.min(dataArray,index1)
                pp.min(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.min(dims)

    def cumprod(self, dataArray, index=None, keep=None):
        """Apply 'cumprod' along some dimension of DataArray.
            Ex:
                pp.cumprod(dataArray)
                pp.cumprod(dataArray,index1)
                pp.cumprod(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.cumprod(dims)

    def cumsum(self, dataArray, index=None, keep=None):
        """Apply 'cumsum' along some dimension of DataArray.
            Ex:
                pp.cumsum(dataArray)
                pp.cumsum(dataArray,index1)
                pp.cumsum(dataArray,[index1,index2]]
        """
        dims = self._getDimsToOperate(dataArray, index, keep)
        return dataArray.cumsum(dims)

    def linearDepreciation(self, investments, usefulLife, timeIndex, includeInCurrentMonth=False):
        """ TODO: comment
            Ex.
                pp.linearDepreciation(investments, usefulLife)
        """
        timeCoords = timeIndex.dataArray
        _cuota = investments/usefulLife/12
        _newTime = xr.DataArray(
            timeCoords.values, [('new_time', timeCoords.values)])
        _ending = self.apply(timeCoords, self.addPeriods, usefulLife*12)
        if includeInCurrentMonth:
            _proy = xr.where((_newTime >= timeCoords) &
                             (_newTime < _ending), _cuota, 0)
        else:
            _proy = xr.where((_newTime > timeCoords) &
                             (_newTime <= _ending), _cuota, 0)
        _final = _proy.sum(timeIndex.name)
        _final = _final.rename({"new_time": timeIndex.name})
        return _final

    def isLinux(self):
        if platform == "linux" or platform == "linux2" or platform == "darwin":
            return True
        else:
            return False


class XIndex(object):
    """ Class for quickly create a xarray coord 
    """

    def __init__(self, name, values, dtype=None, parent=None, mapArrayId=None):
        np_values = np.asarray(values, dtype)
        seripandas = pd.Series(np_values)
        self.name = name
        self.values = seripandas.unique()
        self.mapArrayId = mapArrayId
        self.parent = parent

        """        
            seripandas = pd.Series(np_values[:,0])
            nLevels = np_values.shape[1]
            if levels is None or nLevels!=len(levels):
                raise ValueError("the size of the 'levels' param must be equal to the size of the dimension 2 of the array")
            nn=0
            for levelName in levels:
                levelDic[levelName] = np_values[:,nn]
                nn+=1
        """

    @property
    def coord(self):
        return (self.name, self.values)

    @property
    def dataArray(self):
        tmp = XHelpers(None)
        return tmp.dataArray([self], self.values)

    @property
    def pos(self):
        tmp = XHelpers(None)
        return tmp.dataArray([self], range(0, len(self.values)))

    def __len__(self):
        return len(self.values)

    def __repr__(self):
        return "coord({}, {})".format(self.name, self.values)

    def __eq__(self, other):
        return self.apply_op(self, other, np.equal)

    # A != B
    def __ne__(self, other):
        return self.apply_op(self, other, np.not_equal)

    # A < B
    def __lt__(self, other):
        return self.apply_op(self, other, np.less)

    # A <= B
    def __le__(self, other):
        return self.apply_op(self, other, np.less_equal)

    # A > B
    def __gt__(self, other):
        return self.apply_op(self, other, np.greater)

    # A >= B
    def __ge__(self, other):
        return self.apply_op(self, other, np.greater_equal)

    def apply_op(self, a, b, func, *args):
        if isinstance(a, XIndex):
            a = a.dataArray
        if isinstance(b, XIndex):
            b = b.dataArray

        return func(a, b, *args)
